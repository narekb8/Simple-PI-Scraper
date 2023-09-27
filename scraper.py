# Import libraries
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import sys
import re
from PyPDF2 import PdfReader

# Regex alphabet - RegEx gets complicated really fast, if you need a breakdown of what a regex
# string does, you can use regexr.com and it will lay it out in a much more parsable way
alphabet = r'[^ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz]'
invAlphabet = r'^[^ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz]*'

# Get number of args from cli - argv[0] is always the program name
listCount = int(sys.argv[1])
piRow = str(sys.argv[2])

# Offset for each additional indication added for a drug
indicOffset = 0

# Function to grab the list of drugs from the active worksheet (Cells A2 -> A[listCount])
# To adjust, change min_row to change the starting row, and change min_col and max_col to the number equivalent of the column (A=1, Z=26, AA=27, etc.)
# min_col and max_col SHOULD ALWAYS MATCH
def getList():
    cells = []
    for row in ws.iter_rows(min_row=2, max_row=listCount, min_col=1, max_col=1, values_only=True):
        for cell in row:
            val = str(cell)
            if " " in val:
                val = val.replace(" ", "-") # Replace spaces with a hyphen, even if it doesnt work better than nothing
            cells.append(val)
    
    return cells

# Grab the workbook to use, then set the active sheet as the add list
wkbkpath = <YOUR EXCEL PATH HERE>
wkbk = load_workbook(wkbkpath)
ws = wkbk["Add list"]
template = wkbk["Template"]

# Get list of drug names, then iterate over them 1 by 1
drugList = getList()

for drug in drugList:

    # Path contains local copy of PI, currLink holds the web link to the PI
    path = None
    currLink = None

    # URL from which pdfs to be downloaded
    url = "https://www."+str(drug)+".com"
    # Append drug name to excel sheet
    template['A'+str(drugList.index(drug)+2+indicOffset)] = str(drug)
 
    # Requests URL and get response object, uses a header agent to act as a true browser client
    # Surrounded in try-catch since if the url doesn't exist then the request will simply error out
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:50.0) Gecko/20100101 Firefox/50.0'}
    try:
        response = requests.get(url, timeout=10, headers=header) # 10 seconds
    except:
        print("Timed out " + str(drug))
        continue
    print(str(drug))
    
    # Parse HTML obtained
    soup = BeautifulSoup(response.text, 'html.parser')
    
    # Find all hyperlinks present on webpage
    links = soup.find_all('a')
    
    # From all links check for pdf link
    # If present download file and write link back into worksheet (WIP)
    for link in links:
        if('Prescribing Information' in link.text):
            print(str(drug))
            print(url)
            if (link.get('href', '').endswith('.pdf')  or link.get('href', '').endswith('.ashx')):
                # Get response object for link
                if "http" in link.get('href'):
                    response = requests.get(link.get('href'))
                    currLink = link.get('href', [])
                else:
                    response = requests.get(url+link.get('href'))
                    currLink = url + link.get('href', [])
            
                # Write content in pdf file
                pdf = open("pi-"+drug+".pdf", 'wb')
                pdf.write(response.content)
                path = pdf.name
                pdf.close()
                print("File downloaded")
                break

    # If unable to locate a PI on the main website, attempt to just generically go to the healthcare provider site
    # Rescan from there just like above
    if path is None:
        url = "https://www."+str(drug)+"hcp.com"

        try:
            response = requests.get(url, timeout=10, headers=header) # 10 seconds
        except:
            print("Timed out")
            continue

        soup = BeautifulSoup(response.text, 'html.parser')
        links = soup.find_all('a')

        for link in links:
            if('Prescribing Information' in link.text):
                print(str(drug))
                print(url)
                if (link.get('href', '').endswith('.pdf') or link.get('href', '').endswith('.ashx')):
                    if "http" in link.get('href'):
                        response = requests.get(link.get('href'))
                        currLink = link.get('href', [])
                    else:
                        response = requests.get(url+link.get('href'))
                        currLink = url + link.get('href', [])
            
                    pdf = open("pi-"+drug+".pdf", 'wb')
                    pdf.write(response.content)
                    path = pdf.name
                    pdf.close()
                    print("File downloaded")
                    break
    
    # If we found a PI in either of the queries, download it and parse the file for indications
    if path:
        # Load pdf to parse
        reader = PdfReader(path)
        page = reader.pages[0]
        text = page.extract_text()

        # This is the most forbidden code I have ever written
        # We find the section of the pdf that is between the first instance of indications and a lower boundary (dosage/warnings)
        findex = text.find("INDICATIONS AND USAGE") if text.find("INDICATIONS AND USAGE") != -1 else text.find("INDICATION")
        bindex = text.find("DOSAGE") if text.find("DOSAGE") != -1 else text.find("WARNINGS AND PRECAUTIONS")
        fullIndic = text[text.find('\n', findex)+1:text.rfind('\n', 0, bindex)]

        # Even more forbidden, this string of regex will locate most numerical representations (whole, decimal, fraction) within parentheses
        # Uses this to split the individual indications (still needs to be grouped properly)
        # This is where regexr.com will do you wonders in trying to understand
        indicList = []
        for indic in re.split(r'(\((?<!\d)1(?!\d)+(\.\d)?)', fullIndic):
            if indic is None:
                continue
            elif '(' in indic and ')' not in indic:
                indicList.append(str(indic)+')')
            elif len(str(indic)) < 10:
                continue
            else:
                indicList.append(re.split(invAlphabet, indic, 1)[1])
        
        # Group the indications together properly and separate them from each other to be listed in separate groups
        # Uses two separate lists (could be a map im lazy), one as an index table for indications, the other containing ANOTHER list of all the actual text
        # 
        # Example:
        # [indic0, indic1, indic2]
        # 
        # [[indic0Text0, indic0Text1], [indic1Text0], [indic2Text0, indic2Text1, indic2Text2, indic2Text3]]
        #
        diffIndics = []
        indicMap = []
        j = 0
        while j < len(indicList):
            j = j + 1
            if j < len(indicList):
                if indicList[j] not in diffIndics:
                    diffIndics.append(indicList[j])
                    indicMap.append([indicList[j-1]])
                else:
                    indicMap[diffIndics.index(indicList[j])].append(indicList[j-1])
            j += 1
        
        # If we have less than 20 indices in the list it means we haven't accidentally scraped an entire PI
        # Loop through the index map, combining individual indications into a single string, and append them to the excel sheet
        if len(diffIndics) < 21:
            singleIndic = ""
            j = 0
            while j < len(diffIndics):
                for indic in indicMap[j]:
                    if singleIndic == "":
                        singleIndic = indic
                    else:
                        singleIndic = singleIndic + " " + indic
                print(diffIndics[j])

                template['A'+str(drugList.index(drug)+2+j+indicOffset)] = str(drug)
                template['AF'+str(drugList.index(drug)+2+j+indicOffset)] = str(currLink)
                if template[piRow+str(drugList.index(drug)+2+j+indicOffset)].value is None:
                    template[piRow+str(drugList.index(drug)+2+j+indicOffset)] = str(singleIndic)
                else:
                    template[piRow+str(drugList.index(drug)+2+j+indicOffset)] = str(ws[piRow+str(drugList.index(drug)+1+j)].value) + ' ' + str(singleIndic)
                wkbk.save(wkbkpath)

                singleIndic = ""
                j += 1
            indicOffset += j-1